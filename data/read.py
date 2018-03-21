#coding=utf-8
from openpyxl import load_workbook,Workbook
import logging

logging.basicConfig(filename='c:\\robot\\report\\5.txt',level=logging.DEBUG)
a=[]   
x=load_workbook("1.xlsx")
logging.info("开始读取表格")
y=x["pan"]
logging.info('得到第一个表')
for i in y.rows:
	for j in i:
		a.append(j.value)
print (a)
logging.info('打印数据')
'''
wb=Workbook()
x=wb.active
x1=wb.create_sheet()
x2=wb.create_sheet()
x.title='quapn'
x1.title="meng"
x2.title="panmeng"

for i in range(len(y)):
	x["B%d"%(i+1)]=y[i]
wb.save("3.xlsx")
'''  